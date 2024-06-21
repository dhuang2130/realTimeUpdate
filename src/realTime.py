import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy
import re
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import time
import os

def parse_purchase(purchase, product_key):
    if isinstance(purchase, float) and pd.isna(purchase):
        return []
    purchase_str = str(purchase)
    purchase_str = purchase_str.replace('Ã—', '').replace(',', '')
    items = re.findall(r'(\d+)\s([A-Z0-9-]+|\w+.*)', purchase_str)
    parsed_items = []
    for item in items:
        code = item[1]
        if code in product_key:
            code = product_key[code]
        parsed_items.append((code, int(item[0])))
    return parsed_items

def update_tracking_file(sales_record_file_path, tracking_file_path):
    print("Reading tracking file...")
    if tracking_file_path.endswith('.xls'):
        tracking_df = pd.read_excel(tracking_file_path, sheet_name='Sheet1', engine='xlrd')
    else:
        tracking_df = pd.read_excel(tracking_file_path, sheet_name='Sheet1')

    product_codes = tracking_df.iloc[:, 0].dropna().tolist()  # Assume the first column is the product code

    product_quantities = {code: 0 for code in product_codes}

    print("Reading product key from stats tab...")
    if sales_record_file_path.endswith('.xls'):
        stats_df = pd.read_excel(sales_record_file_path, sheet_name='Stats', usecols=[1, 2], engine='xlrd')
    else:
        stats_df = pd.read_excel(sales_record_file_path, sheet_name='Stats', usecols=[1, 2])

    product_key = dict(zip(stats_df.iloc[:, 1], stats_df.iloc[:, 0]))

    print("Reading sales record file...")
    if sales_record_file_path.endswith('.xls'):
        sales_record_df = pd.read_excel(sales_record_file_path, sheet_name='Raw', engine='xlrd')
    else:
        sales_record_df = pd.read_excel(sales_record_file_path, sheet_name='Raw')

    print("Summing quantities for each product...")
    for purchase in sales_record_df['Purchase']:
        items = parse_purchase(purchase, product_key)
        for item, quantity in items:
            if item in product_quantities:
                product_quantities[item] += quantity

    print("Updating tracking file with new quantities...")
    if tracking_file_path.endswith('.xls'):
        rb = open_workbook(tracking_file_path, formatting_info=True)
        wb = copy(rb)
        ws = wb.get_sheet(0)
        for idx, code in enumerate(product_codes, start=1):
            if code in product_quantities:
                ws.write(idx, 3, product_quantities[code])
        wb.save(tracking_file_path)
    else:
        workbook = load_workbook(tracking_file_path)
        sheet = workbook['Sheet1']

        sales_column_index = 4  
        for idx, code in enumerate(product_codes, start=2):
            if code in product_quantities:
                sheet.cell(row=idx, column=sales_column_index).value = product_quantities[code]

        # Save the workbook
        workbook.save(tracking_file_path)

    print("The quantities have been updated in the existing 'Sales' column of the tracking file.")

class SalesRecordHandler(FileSystemEventHandler):
    def __init__(self, sales_record_file_path, tracking_file_path):
        self.sales_record_file_path = sales_record_file_path
        self.tracking_file_path = tracking_file_path

    def on_modified(self, event):
        print(f"Detected change in: {event.src_path}")
        if os.path.abspath(event.src_path) == os.path.abspath(self.sales_record_file_path):
            print("Updating tracking file...")
            try:
                update_tracking_file(self.sales_record_file_path, self.tracking_file_path)
            except PermissionError as e:
                print(f"Permission error: {e}")

if __name__ == "__main__":
    sales_record_file_path = './data/Z2024SalesRecord.xlsx'
    tracking_file_path = './data/realtime Tracking.xlsx'
    
    try:
        update_tracking_file(sales_record_file_path, tracking_file_path)
        print("Running initial update...")
        event_handler = SalesRecordHandler(sales_record_file_path, tracking_file_path)
        observer = Observer()
        observer.schedule(event_handler, path=os.path.dirname(sales_record_file_path), recursive=False)
        observer.start()
        print("Started monitoring for changes...")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            observer.stop()
            observer.join()
            print("Stopped monitoring for changes.")
    except PermissionError as e:
        print(f"Permission error: {e}")
